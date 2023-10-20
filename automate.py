import time
import pyautogui
import tkinter as tk
from tkinter import simpledialog, filedialog
from docx import Document
import openpyxl
import pyperclip


def get_excel_file():
    #Prompt the user to select an Excel file and return the file path.
    return filedialog.askopenfilename(title="Select Excel Document", filetypes=[("Excel Files", "*.xlsx")])

def get_word_file():
    #Prompt the user to select a Word file and return the file path.
    return filedialog.askopenfilename(title="Select Word Document", filetypes=[("Word Files", "*.docx")])

def extract_content_from_word(doc_path):
    #Extract and return the text content from the specified Word document.
    doc = Document(doc_path)
    return '\n'.join(paragraph.text for paragraph in doc.paragraphs)

def send_email_via_outlook(email_address, subject_line, content):
    #Automate the process of opening Outlook and sending an email with the provided details.
    pyautogui.press("win")
    pyautogui.write("Outlook")
    pyautogui.press("enter")
    time.sleep(2.7)
    pyautogui.hotkey("ctrl", "n")
    time.sleep(0.9)
    pyautogui.write(email_address)
    time.sleep(0.9)
    for _ in range(3):
        pyautogui.press("tab")
        time.sleep(0.4)
    pyautogui.write(subject_line)
    time.sleep(0.3)
    pyautogui.press("tab")
    pyperclip.copy(content)
    pyautogui.hotkey("ctrl", "v")

def main():
    #Main function to execute the email automation program.
    root = tk.Tk()
    root.withdraw()
    excel_file = get_excel_file()
    if not excel_file:
        print("No Excel document selected. Exiting.")
        return
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    id_email = simpledialog.askinteger("Subject Input", "Enter the user ID:")
    if not id_email:
        print("No user ID entered. Exiting.")
        return
    email_address = sheet.cell(row=id_email, column=3).value
    user_name = sheet.cell(row=id_email, column=1).value
    if not email_address:
        print("No email address found in the Excel document. Exiting.")
        return
    subject_line = simpledialog.askstring("Subject Line Input", "Enter the Subject Line:")
    word_file = get_word_file()
    if not word_file:
        print("No Word document selected. Exiting.")
        return
    text_content = extract_content_from_word(word_file)
    email_content = text_content.replace('{name}', user_name)
    send_email_via_outlook(email_address, subject_line, email_content)

if __name__ == '__main__':
    main()